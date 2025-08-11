"""
报告导出工具
支持PDF、Excel、JSON格式的报告导出功能
"""

import json
import pandas as pd
from typing import Dict, Any, Optional, List
from datetime import datetime
from pathlib import Path
import logging
import io
import base64

logger = logging.getLogger(__name__)

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("ReportLab not available. PDF export will be disabled.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("OpenPyXL not available. Excel export will be disabled.")


class ReportExporter:
    """报告导出器"""
    
    def __init__(self):
        self.supported_formats = ['json']
        if REPORTLAB_AVAILABLE:
            self.supported_formats.append('pdf')
        if OPENPYXL_AVAILABLE:
            self.supported_formats.append('excel')
    
    def export_report(self, report_data: Dict[str, Any], format_type: str, 
                     output_path: Optional[str] = None) -> Dict[str, Any]:
        """
        导出报告
        
        Args:
            report_data: 报告数据
            format_type: 导出格式 ('pdf', 'excel', 'json')
            output_path: 输出文件路径
            
        Returns:
            导出结果
        """
        try:
            format_type = format_type.lower()
            
            if format_type not in self.supported_formats:
                return {
                    'status': 'error',
                    'message': f'不支持的导出格式: {format_type}。支持的格式: {", ".join(self.supported_formats)}'
                }
            
            # 生成默认文件名
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"reports/analysis_report_{timestamp}.{format_type}"
            
            # 确保输出目录存在
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # 根据格式调用相应的导出方法
            if format_type == 'json':
                result = self._export_json(report_data, output_path)
            elif format_type == 'pdf':
                result = self._export_pdf(report_data, output_path)
            elif format_type == 'excel':
                result = self._export_excel(report_data, output_path)
            else:
                return {
                    'status': 'error',
                    'message': f'未实现的导出格式: {format_type}'
                }
            
            if result['status'] == 'success':
                result['file_path'] = output_path
                result['file_size'] = Path(output_path).stat().st_size
                
            return result
            
        except Exception as e:
            logger.error(f"报告导出失败: {e}")
            return {
                'status': 'error',
                'message': f'导出失败: {str(e)}'
            }
    
    def _export_json(self, report_data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """导出JSON格式报告"""
        try:
            # 添加导出元数据
            export_data = {
                'export_metadata': {
                    'format': 'json',
                    'exported_at': datetime.now().isoformat(),
                    'version': '1.0'
                },
                'report_data': report_data
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
            
            return {
                'status': 'success',
                'message': 'JSON报告导出成功'
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'JSON导出失败: {str(e)}'
            }
    
    def _export_pdf(self, report_data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """导出PDF格式报告"""
        if not REPORTLAB_AVAILABLE:
            return {
                'status': 'error',
                'message': 'PDF导出功能不可用，请安装reportlab库'
            }
        
        try:
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # 标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # 居中
            )
            story.append(Paragraph("用户行为分析报告", title_style))
            story.append(Spacer(1, 20))
            
            # 报告元数据
            metadata = report_data.get('metadata', {})
            if metadata:
                story.append(Paragraph("报告信息", styles['Heading2']))
                
                meta_data = [
                    ['生成时间', metadata.get('generated_at', 'N/A')],
                    ['数据范围', str(metadata.get('analysis_scope', {}).get('date_range', 'All time'))],
                ]
                
                data_summary = metadata.get('data_summary', {})
                if data_summary:
                    meta_data.extend([
                        ['总事件数', f"{data_summary.get('total_events', 0):,}"],
                        ['独立用户数', f"{data_summary.get('unique_users', 0):,}"],
                        ['总会话数', f"{data_summary.get('total_sessions', 0):,}"]
                    ])
                
                meta_table = Table(meta_data)
                meta_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(meta_table)
                story.append(Spacer(1, 20))
            
            # 执行摘要
            executive_summary = report_data.get('executive_summary', {})
            if executive_summary:
                story.append(Paragraph("执行摘要", styles['Heading2']))
                
                key_metrics = executive_summary.get('key_metrics', {})
                if key_metrics:
                    story.append(Paragraph("关键指标", styles['Heading3']))
                    for metric, value in key_metrics.items():
                        if isinstance(value, (int, float)):
                            if metric.endswith('_rate'):
                                story.append(Paragraph(f"• {metric}: {value:.2%}", styles['Normal']))
                            else:
                                story.append(Paragraph(f"• {metric}: {value:,}", styles['Normal']))
                        else:
                            story.append(Paragraph(f"• {metric}: {value}", styles['Normal']))
                    story.append(Spacer(1, 12))
                
                # 关键趋势
                key_trends = executive_summary.get('key_trends', [])
                if key_trends:
                    story.append(Paragraph("关键趋势", styles['Heading3']))
                    for trend in key_trends:
                        story.append(Paragraph(f"• {trend}", styles['Normal']))
                    story.append(Spacer(1, 12))
                
                # 关键问题
                key_issues = executive_summary.get('key_issues', [])
                if key_issues:
                    story.append(Paragraph("关键问题", styles['Heading3']))
                    for issue in key_issues:
                        story.append(Paragraph(f"• {issue}", styles['Normal']))
                    story.append(Spacer(1, 12))
            
            # 详细分析结果
            detailed_analysis = report_data.get('detailed_analysis', {})
            if detailed_analysis:
                story.append(PageBreak())
                story.append(Paragraph("详细分析", styles['Heading2']))
                
                for analysis_type, analysis_data in detailed_analysis.items():
                    if isinstance(analysis_data, dict) and 'error' not in analysis_data:
                        story.append(Paragraph(analysis_type.replace('_', ' ').title(), styles['Heading3']))
                        
                        # 添加摘要信息
                        summary = analysis_data.get('summary', {})
                        if summary:
                            for key, value in summary.items():
                                if isinstance(value, (int, float)):
                                    if key.endswith('_rate'):
                                        story.append(Paragraph(f"• {key}: {value:.2%}", styles['Normal']))
                                    else:
                                        story.append(Paragraph(f"• {key}: {value:,}", styles['Normal']))
                                elif isinstance(value, list):
                                    story.append(Paragraph(f"• {key}: {', '.join(map(str, value[:3]))}", styles['Normal']))
                                else:
                                    story.append(Paragraph(f"• {key}: {value}", styles['Normal']))
                        
                        story.append(Spacer(1, 12))
            
            # 构建PDF
            doc.build(story)
            
            return {
                'status': 'success',
                'message': 'PDF报告导出成功'
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'PDF导出失败: {str(e)}'
            }
    
    def _export_excel(self, report_data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """导出Excel格式报告"""
        if not OPENPYXL_AVAILABLE:
            return {
                'status': 'error',
                'message': 'Excel导出功能不可用，请安装openpyxl库'
            }
        
        try:
            workbook = openpyxl.Workbook()
            
            # 删除默认工作表
            workbook.remove(workbook.active)
            
            # 创建摘要工作表
            summary_sheet = workbook.create_sheet("报告摘要")
            self._create_summary_sheet(summary_sheet, report_data)
            
            # 创建详细分析工作表
            detailed_analysis = report_data.get('detailed_analysis', {})
            for analysis_type, analysis_data in detailed_analysis.items():
                if isinstance(analysis_data, dict) and 'error' not in analysis_data:
                    sheet_name = analysis_type.replace('_', ' ').title()[:31]  # Excel工作表名称限制
                    sheet = workbook.create_sheet(sheet_name)
                    self._create_analysis_sheet(sheet, analysis_type, analysis_data)
            
            # 保存工作簿
            workbook.save(output_path)
            
            return {
                'status': 'success',
                'message': 'Excel报告导出成功'
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Excel导出失败: {str(e)}'
            }
    
    def _create_summary_sheet(self, sheet, report_data: Dict[str, Any]):
        """创建Excel摘要工作表"""
        # 设置标题
        sheet['A1'] = "用户行为分析报告摘要"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:D1')
        
        row = 3
        
        # 报告元数据
        metadata = report_data.get('metadata', {})
        if metadata:
            sheet[f'A{row}'] = "报告信息"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            sheet[f'A{row}'] = "生成时间"
            sheet[f'B{row}'] = metadata.get('generated_at', 'N/A')
            row += 1
            
            data_summary = metadata.get('data_summary', {})
            if data_summary:
                for key, value in data_summary.items():
                    sheet[f'A{row}'] = key.replace('_', ' ').title()
                    sheet[f'B{row}'] = value
                    row += 1
            
            row += 1
        
        # 执行摘要
        executive_summary = report_data.get('executive_summary', {})
        if executive_summary:
            sheet[f'A{row}'] = "关键指标"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            key_metrics = executive_summary.get('key_metrics', {})
            for metric, value in key_metrics.items():
                sheet[f'A{row}'] = metric.replace('_', ' ').title()
                sheet[f'B{row}'] = value
                row += 1
    
    def _create_analysis_sheet(self, sheet, analysis_type: str, analysis_data: Dict[str, Any]):
        """创建Excel分析工作表"""
        # 设置标题
        sheet['A1'] = analysis_type.replace('_', ' ').title()
        sheet['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # 添加摘要信息
        summary = analysis_data.get('summary', {})
        if summary:
            sheet[f'A{row}'] = "摘要"
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for key, value in summary.items():
                sheet[f'A{row}'] = key.replace('_', ' ').title()
                if isinstance(value, list):
                    sheet[f'B{row}'] = ', '.join(map(str, value[:5]))
                else:
                    sheet[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # 添加详细数据（如果有DataFrame格式的数据）
        for key, value in analysis_data.items():
            if key != 'summary' and isinstance(value, dict):
                sheet[f'A{row}'] = key.replace('_', ' ').title()
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                # 如果是字典，展示键值对
                for sub_key, sub_value in value.items():
                    sheet[f'A{row}'] = f"  {sub_key}"
                    sheet[f'B{row}'] = str(sub_value)[:100]  # 限制长度
                    row += 1
                
                row += 1
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的导出格式"""
        return self.supported_formats.copy()
    
    def validate_report_data(self, report_data: Dict[str, Any]) -> Dict[str, Any]:
        """验证报告数据格式"""
        try:
            required_fields = ['metadata', 'detailed_analysis']
            missing_fields = []
            
            for field in required_fields:
                if field not in report_data:
                    missing_fields.append(field)
            
            if missing_fields:
                return {
                    'valid': False,
                    'message': f'缺少必需字段: {", ".join(missing_fields)}'
                }
            
            return {
                'valid': True,
                'message': '报告数据格式有效'
            }
            
        except Exception as e:
            return {
                'valid': False,
                'message': f'数据验证失败: {str(e)}'
            }


def create_download_link(file_path: str, link_text: str) -> str:
    """创建文件下载链接（用于Streamlit）"""
    try:
        with open(file_path, 'rb') as f:
            file_data = f.read()
        
        b64_data = base64.b64encode(file_data).decode()
        file_name = Path(file_path).name
        
        return f'<a href="data:application/octet-stream;base64,{b64_data}" download="{file_name}">{link_text}</a>'
    
    except Exception as e:
        logger.error(f"创建下载链接失败: {e}")
        return f"下载链接创建失败: {str(e)}"